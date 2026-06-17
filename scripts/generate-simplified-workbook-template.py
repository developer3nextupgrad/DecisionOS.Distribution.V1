#!/usr/bin/env python3
"""Generate DecisionOS simplified multi-tab workbook template for best dashboard coverage."""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "src" / "DecisionOS.Distribution.Web" / "wwwroot" / "downloads" / "DecisionOS_Simplified_Workbook_Template.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def style_header(ws, ncol: int) -> None:
    for col in range(1, ncol + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for col in range(1, ncol + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def week_ends(start: date, count: int) -> list[date]:
    # Saturdays
    d = start
    while d.weekday() != 5:
        d += timedelta(days=1)
    return [d + timedelta(weeks=i) for i in range(count)]


def build_readme(ws) -> None:
    rows = [
        ["Topic", "Guidance"],
        ["Purpose", "Fill this workbook and upload via Operations → Uploads → Simplified."],
        ["Weekly_Financials", "Required for all 7 dashboard KPIs. Keep Week_End_Date on every row."],
        ["Net_Profit_%", "Ratio (0.06) or percent (6). Required for Net Profit KPI — not derivable from gross margin alone."],
        ["AR_Ending / AP_Ending", "Week-ending dollar balances — used for Cash Conversion Cycle (CCC)."],
        ["AR_Over_60_%", "Past-due AR ratio/percent. Maps to AR Health KPI (31+ approximation)."],
        ["Sales / AR / AP tabs", "Detail rows optional but improve driver accuracy; AR/AP use Open_Amount + Age_Days."],
        ["Extra tabs", "Set role Skip on README and any unused tabs during Import Review."],
    ]
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    style_header(ws, 2)


def build_weekly_financials(ws, weeks: list[date]) -> None:
    headers = [
        "Week_Number", "Week_End_Date", "Net_Sales", "COGS", "Gross_Profit", "Gross_Margin_%",
        "Orders", "Active_Customers", "Avg_Order_Value", "Inventory_Value_End",
        "Fill_Rate_%", "AR_Ending", "AR_Over_60_%", "AP_Ending", "AP_Past_Due_%",
        "Net_Profit_%", "Net_Income", "Cash_Ending", "Notes",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, len(headers))

    base_sales = 680_000
    for i, we in enumerate(weeks):
        net = base_sales + i * 12_000
        cogs = round(net * 0.72, 2)
        gp = net - cogs
        gm = round(gp / net, 4)
        np_pct = 0.058 + i * 0.001
        net_income = round(net * np_pct, 2)
        ws.append([
            i + 1, we.isoformat(), net, cogs, gp, gm,
            420 + i * 5, 118 + i, round(net / (420 + i * 5), 2),
            1_050_000 + i * 8_000, 0.94 - i * 0.005,
            820_000 + i * 10_000, 0.14 - i * 0.005,
            310_000 + i * 4_000, 0.11 - i * 0.003,
            np_pct, net_income, 240_000 + i * 5_000, "Template sample week",
        ])


def build_sales(ws, weeks: list[date]) -> None:
    headers = [
        "Week_Number", "Week_End_Date", "SKU", "Category", "Customer_ID", "Units_Sold",
        "Gross_Sales", "Discount_Amount", "Net_Sales", "COGS", "Gross_Profit", "Gross_Margin_%",
        "Channel", "Order_Count",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, len(headers))

    skus = [("SKU-100", "Balls"), ("SKU-200", "Shoes"), ("SKU-300", "Bags")]
    customers = ["CUST-001", "CUST-002", "CUST-003"]
    row_num = 2
    for i, we in enumerate(weeks):
        for j, (sku, cat) in enumerate(skus):
            net = 40_000 + i * 2_000 + j * 5_000
            cogs = round(net * 0.72, 2)
            gp = net - cogs
            ws.append([
                i + 1, we.isoformat(), sku, cat, customers[j % 3], 80 + j * 10,
                net * 1.02, round(net * 0.02, 2), net, cogs, gp, round(gp / net, 4),
                "Outside", 12 + j,
            ])
            row_num += 1


def build_ar(ws, weeks: list[date]) -> None:
    headers = [
        "Invoice_ID", "Customer_ID", "Customer_Name", "Invoice_Date", "Due_Date",
        "Original_Amount", "Open_Amount", "Days_Past_Due", "Aging_Bucket", "Collection_Status",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, len(headers))

    we = weeks[-1]
    buckets = [("Current", 0), ("1-30", 18), ("31-60", 45), ("61-90", 75), ("90+", 95)]
    for n, (bucket, days) in enumerate(buckets):
        amt = 50_000 + n * 12_000
        inv = we - timedelta(days=days + 10)
        due = inv + timedelta(days=30)
        ws.append([
            f"INV-{1000+n}", f"CUST-00{(n % 3) + 1}", f"Customer {(n % 3) + 1}",
            inv.isoformat(), due.isoformat(), amt, amt, days, bucket, "Open",
        ])


def build_ap(ws, weeks: list[date]) -> None:
    headers = [
        "Bill_ID", "Vendor_ID", "Vendor_Name", "Bill_Date", "Due_Date",
        "Original_Amount", "Open_Amount", "Days_Past_Due", "Aging_Bucket", "Payment_Status",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, len(headers))

    we = weeks[-1]
    for n in range(4):
        amt = 28_000 + n * 9_000
        days = [0, 22, 48, 82][n]
        bill = we - timedelta(days=days + 5)
        due = bill + timedelta(days=30)
        ws.append([
            f"BILL-{2000+n}", f"VEND-00{n+1}", f"Vendor {n+1}",
            bill.isoformat(), due.isoformat(), amt, amt, days,
            ["Current", "1-30", "31-60", "90+"][n], "Open",
        ])


def build_inventory(ws, weeks: list[date]) -> None:
    headers = [
        "Snapshot_Date", "SKU", "Category", "On_Hand_Units", "Unit_Cost", "Inventory_Value",
        "Last_Sale_Date", "Quantity_On_Hand",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, len(headers))

    snap = weeks[-1]
    items = [("SKU-100", 120, 42), ("SKU-200", 85, 78), ("SKU-300", 200, 15)]
    for sku, qty, cost in items:
        val = qty * cost
        ws.append([snap.isoformat(), sku, "Category", qty, cost, val, snap.isoformat(), qty])


def build_customers(ws) -> None:
    headers = [
        "Customer_ID", "Customer_Name", "Customer_Type", "Payment_Terms", "Credit_Limit", "Active_Flag",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, len(headers))
    ws.append(["CUST-001", "Acme Lanes", "Commercial", "Net 30", 50000, "Y"])
    ws.append(["CUST-002", "Pro Shop East", "Commercial", "Net 30", 35000, "Y"])
    ws.append(["CUST-003", "League Central", "League", "Net 15", 20000, "Y"])


def build_vendors(ws) -> None:
    headers = [
        "Vendor_ID", "Vendor_Name", "Payment_Terms", "Lead_Time_Days", "Fill_Rate_%", "On_Time_%",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, len(headers))
    ws.append(["VEND-001", "Strike Supply Co", "Net 30", 7, 0.96, 0.94])
    ws.append(["VEND-002", "Pin Partners", "Net 45", 10, 0.93, 0.91])
    ws.append(["VEND-003", "Lane Logistics", "Net 30", 5, 0.95, 0.97])


def build_holdover(ws) -> None:
    headers = [
        "Holdover_ID", "Customer_ID", "Customer_Name", "Area", "Action", "Owner", "Status", "Completion_%",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, len(headers))
    ws.append(["HO-1", "CUST-001", "Acme Lanes", "AR", "Call on 90+ accounts", "Collections", "Open", 35])


def main() -> None:
    weeks = week_ends(date(2025, 11, 22), 4)
    wb = Workbook()
    wb.remove(wb.active)

    sheets = [
        ("README_Import_Map", lambda ws: build_readme(ws)),
        ("Weekly_Financials", lambda ws: build_weekly_financials(ws, weeks)),
        ("Sales_By_SKU_Week", lambda ws: build_sales(ws, weeks)),
        ("Accounts_Receivable", lambda ws: build_ar(ws, weeks)),
        ("Accounts_Payable", lambda ws: build_ap(ws, weeks)),
        ("Inventory_By_SKU", lambda ws: build_inventory(ws, weeks)),
        ("Customer_Master", lambda ws: build_customers(ws)),
        ("Vendor_Master", lambda ws: build_vendors(ws)),
        ("Holdover_Actions", lambda ws: build_holdover(ws)),
    ]

    for name, builder in sheets:
        ws = wb.create_sheet(name)
        builder(ws)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
